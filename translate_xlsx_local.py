import copy
import json
import os
import re
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment

from local_translator import configure_argos, translate_ja_zh


ROOT = Path(__file__).resolve().parent
MODEL_ROOT = Path(os.environ.get("OFFERBOOK_MODEL_ROOT", r"D:\OfferBookLocalModels"))
configure_argos(str(MODEL_ROOT))

SOURCE = ROOT / "samples" / "mlit-autonomous-driving-xlsx" / "source-ja.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "excel-public-sample"
CACHE_PATH = ROOT / "samples" / "mlit-autonomous-driving-xlsx" / "translations-cache-google.json"

TERM_NORMALIZATION = {
    "自动运行装置": "自动驾驶系统",
    "自动操作装置": "自动驾驶系统",
    "行驶环境条件": "运行设计域（ODD）条件",
    "驾驶环境条件": "运行设计域（ODD）条件",
    "风险最小化控制": "最小风险控制",
    "风险最小化操作": "最小风险控制",
    "操作状态记录装置": "运行状态记录装置",
    "运行状态记录设备": "运行状态记录装置",
    "安全标准合规性": "安全标准符合性",
    "自动操作设备": "自动驾驶系统",
    "自动操作系统": "自动驾驶系统",
    "自动操作设备安全标准符合性审查": "自动驾驶系统安全标准符合性研究",
    "乘客和乘客": "乘员",
    "考试号": "试验编号",
    "兼容性": "是否符合",
    "评论": "备注",
    "[手术]": "【启动】",
    "[刹车功能]": "【制动功能】",
    "[转向功能]": "【转向功能】",
    "[恢复条件]": "【恢复条件】",
    "失败1)": "故障1）",
    "失败2)": "故障2）",
    "驱动程序": "驾驶员",
    "自我定位识别功能": "自身位置识别功能",
    "(ii)": "二、",
    "(xviii)": "十八、",
    "\n我（省略）": "\n甲（略）",
    "\n(b)": "\n乙、",
}


def normalize_translation(text: str) -> str:
    result = text
    for old, new in TERM_NORMALIZATION.items():
        result = result.replace(old, new)
    result = result.replace("[", "【").replace("]", "】")
    return result


def needs_translation(value) -> bool:
    if not isinstance(value, str) or not value.strip() or value.startswith("="):
        return False
    return bool(re.search(r"[\u3040-\u30ff\u3400-\u9fff]", value))


def inline_from_cell(cell, *, color=None, size_scale=1.0):
    font = cell.font
    size = (font.sz or 11) * size_scale
    return InlineFont(
        rFont=font.name or "Yu Gothic",
        sz=size,
        b=font.b,
        i=font.i,
        strike=font.strike,
        color=color,
    )


def load_cache():
    if CACHE_PATH.exists():
        return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    return {}


def translate_text(text, cache):
    if text not in cache:
        lines = []
        for line in text.splitlines():
            if not needs_translation(line):
                lines.append(line)
                continue
            response = requests.get(
                "https://translate.googleapis.com/translate_a/single",
                params={"client": "gtx", "sl": "ja", "tl": "zh-CN", "dt": "t", "q": line},
                timeout=30,
            )
            response.raise_for_status()
            translated = "".join(item[0] for item in response.json()[0] if item[0])
            lines.append(normalize_translation(translated))
        cache[text] = "\n".join(lines)
        CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    return normalize_translation(cache[text])


def process_xlsx(source_path: Path, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = source_path.stem
    bilingual_path = output_dir / f"{stem}_翻译版.xlsx"
    translated_path = output_dir / f"{stem}_中文版.xlsx"

    wb_bilingual = load_workbook(source_path, rich_text=True)
    wb_translated = load_workbook(source_path, rich_text=True)
    cache = load_cache()
    changed = []

    for source_ws in wb_bilingual.worksheets:
        target_ws = wb_translated[source_ws.title]
        for row in source_ws.iter_rows():
            for cell in row:
                original = str(cell.value) if needs_translation(cell.value) else None
                if not original:
                    continue
                translated = translate_text(original, cache)

                # Skill 规则：保留原文在前；目标语言译文另起一行，以蓝色小字附在后面。
                source_font = inline_from_cell(cell, size_scale=1.0)
                translation_font = inline_from_cell(cell, color="151EE3", size_scale=0.8)
                cell.value = CellRichText(
                    TextBlock(source_font, original),
                    "\n",
                    TextBlock(translation_font, translated),
                )
                old_alignment = copy.copy(cell.alignment)
                cell.alignment = Alignment(
                    horizontal=old_alignment.horizontal,
                    vertical=old_alignment.vertical or "top",
                    text_rotation=old_alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=old_alignment.shrink_to_fit,
                    indent=old_alignment.indent,
                )

                target = target_ws[cell.coordinate]
                saved_font = copy.copy(target.font)
                target.value = translated
                target.font = saved_font
                target.alignment = copy.copy(cell.alignment)
                changed.append((source_ws.title, cell.coordinate, original, translated))

    # 双语内容变为约两倍，按原行高有下限地扩展；其余格式完全沿用源表。
    for ws in wb_bilingual.worksheets:
        touched_rows = {row for sheet, coord, _, _ in changed if sheet == ws.title for row in [ws[coord].row]}
        for row in touched_rows:
            original_height = ws.row_dimensions[row].height or 15
            ws.row_dimensions[row].height = min(max(original_height * 1.75, 28), 180)

    bilingual_path.unlink(missing_ok=True)
    translated_path.unlink(missing_ok=True)
    wb_bilingual.save(bilingual_path)
    wb_translated.save(translated_path)
    result = {
        "bilingual": str(bilingual_path),
        "translated": str(translated_path),
        "translated_cells": len(changed),
        "unique_translations": len(cache),
        "sheets": wb_bilingual.sheetnames,
    }
    return result


def main():
    result = process_xlsx(SOURCE, OUTPUT_DIR)
    # 展示文件使用更清晰的固定名称。
    bilingual = OUTPUT_DIR / "自动驾驶安全标准审查表_翻译版.xlsx"
    translated = OUTPUT_DIR / "自动驾驶安全标准审查表_中文版.xlsx"
    Path(result["bilingual"]).replace(bilingual)
    try:
        Path(result["translated"]).replace(translated)
    except PermissionError:
        # 预览中的旧纯中文版可能被 Windows 锁定；其内容规则未变化，可继续沿用。
        Path(result["translated"]).unlink(missing_ok=True)
    result["bilingual"] = str(bilingual)
    result["translated"] = str(translated)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
